# A small script to file that takes the registered subjects
# and puts them in the .xlsx spreasheet template.
# P.S: YOU NEED TO INSTALL openpyxl module for it to work!
# enter this in your terminal: pip install openpyxl
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

with open("output.csv") as output_table:
    wb = load_workbook(filename = 'table.xlsx')
    ws = wb.active
    ws.title = "table"

    output_table = csv.reader(output_table, delimiter=',')

    # Find another way to acess last row.
    row = output_table
    for row in output_table:
        # Similarly here, Find another way to deal with last row
        if(len(row) == 1): 
            break

        days = int(row[4])
        period_start = int(row[5][0])
        period_end = int(row[5][2])

        #This should be changed to the preferred format by the module.
        for i in range(period_start, period_end + 1):
            ws[f"{chr(i + 65)}{days + 2}"].font = Font(size = 10)
            ws[f"{chr(i + 65)}{days + 2}"] = f"{row[0]} \n- {row[1]} - G{row[3]} \n- Units: {row[2]}"
            ws[f"{chr(i + 65)}{days + 2}"].alignment = Alignment(wrap_text = True, horizontal= 'center', vertical= 'center')
            
    ws['B9'] = row[-1][0]
    ws['B9'].font = Font(size = 10)
    ws['B9'].alignment = Alignment(wrap_text = True, horizontal= 'center', vertical= 'center')
    wb.save('table.xlsx')
